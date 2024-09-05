from flask import Flask, render_template, Response, redirect, url_for
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Load known faces and their names
known_face_encodings = []
known_face_names = []

# Load images from the 'photos' directory
known_faces_dir = 'photos'
for filename in os.listdir(known_faces_dir):
    if filename.endswith(".jpg") or filename.endswith(".png"):
        img = face_recognition.load_image_file(os.path.join(known_faces_dir, filename))
        encoding = face_recognition.face_encodings(img)[0]
        known_face_encodings.append(encoding)
        known_face_names.append(os.path.splitext(filename)[0])

# Function to mark attendance in an Excel sheet
def mark_attendance(name):
    now = datetime.now()
    date_string = now.strftime('%Y-%m-%d')
    time_string = now.strftime('%H:%M:%S')

    # Load or create the attendance Excel sheet
    file_name = 'Attendance.xlsx'
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Date", "Time"])  # Add headers if the file doesn't exist

    # Check if the student is already marked as present
    already_present = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == date_string:
            already_present = True
            break

    # If not already present, mark attendance
    if not already_present:
        sheet.append([name, date_string, time_string])
        workbook.save(file_name)

# Function to generate frames for the webcam
def gen_frames():
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # Resize frame for faster processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        # Find all faces and face encodings in the current frame
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        # Loop through each face found in the frame
        for face_encoding, face_location in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)

            if matches[best_match_index]:
                name = known_face_names[best_match_index]
                mark_attendance(name)

                # Draw a box around the face and label the name
                top, right, bottom, left = [v * 4 for v in face_location]
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
                cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 255), 2)

        # Encode the frame in JPEG format
        ret, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()

        # Yield frame in bytes format
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

    cap.release()

# Route for the dashboard
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

# Route for video streaming
@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

# New Route for Attendance Page
@app.route('/attendance/<class_name>')
def attendance(class_name):
    current_date = datetime.now().strftime('%Y-%m-%d')
    return render_template('attendance.html', current_date=current_date)
    #return render_template('attendance.html', class_name=class_name)

if __name__ == '__main__':
    app.run(debug=True)
